import os
import openpyxl
 
class Exel():

	def __init__(self):
		self.currentPath = os.getcwd()

		self.write_wb = openpyxl.load_workbook(self.currentPath + '/TC.xlsx')
		self.write_ws = self.write_wb.active
	
	def loadSubject(self, num):
		print('\n%d. ' %num + self.write_ws.cell(row=1+num,column=2).value)

	def writeResult(self, num, result):
		if result == True:
			self.write_ws.cell(row=1+num,column=6).value = 'Passed'
		else:
			self.write_ws.cell(row=1+num,column=6).value = 'Failed'			
		self.write_wb.save(self.currentPath + '/TC.xlsx')

